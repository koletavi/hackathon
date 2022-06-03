# coding: utf-8
import threading

import serial  # Serial imported for Serial communication
import time  # Required to use delay functions
import requests
import base64
import json
from tkinter import *
import tkinter as tk
from  tkinter import ttk
from PIL import Image, ImageTk
import openpyxl
from datetime import datetime
import pandas as pd


def excel_read():
    excel_data  = pd.read_excel (r'C:\Users\amrus\Desktop\project\אקתון\sensorData.xlsx')
    time = excel_data["time"].tolist()
    sensor = excel_data["sensor"].tolist()
    monitor = excel_data["monitor"].tolist()
    return time,sensor,monitor
def ux(readData = "123456789"):
    while(True):
        window = tk.Tk()
        greeting = tk.Label(text="Monitoring Insufficiency Of Water Flow In The Sewer",bg="white",font=(None, 16))
        greeting.pack()
        window.configure(background='white')
        window.geometry("1000x500")

        #space
        spacer = tk.Label(text="",font=(None, 40))
        spacer.pack()

        #add image
        # image = Image.open(r"C:\Users\amrus\Desktop\project\אקתון\ramat gan.png")
        # photo = ImageTk.PhotoImage(image)
        # label = Label(image = photo)
        # label.image = photo
        # label.grid(row=1)
        # label.pack()

        #table moitor
        table = Frame(window)
        table.pack()

        table = ttk.Treeview(table)
        table['columns'] = ('EventStartTime', 'SensorNumber', 'StreetsAtRisk', 'SesnorData')
        table.column("#0", width=0,  stretch=NO)
        table.column("EventStartTime",anchor=CENTER, width=120)
        table.column("SensorNumber",anchor=CENTER,width=120)
        table.column("StreetsAtRisk",anchor=CENTER,width=120)
        table.column("SesnorData",anchor=CENTER,width=120)

        table.heading("#0",text="",anchor=CENTER)
        table.heading("EventStartTime",text="Event Start Time",anchor=CENTER)
        table.heading("SensorNumber",text="Sensor Number",anchor=CENTER)
        table.heading("StreetsAtRisk",text="Streets At Risk",anchor=CENTER)
        table.heading("SesnorData",text="Sesnor Data",anchor=CENTER)

        def Refresher(temp =0):
            time,sensor,monitor= excel_read()
            for i in range(1,len(time)):
                # privates_companies = street(data_alert[i][1])
                # if temp != 1:
                #     temp =1
                table.insert(parent='',index='end',iid=i-1,text='',
                values=(time[i],sensor[i],"נקר יוסף, ירושלים",monitor[i]))

            window.after(20000, Refresher)  # every second...
        table.pack()
        Refresher()
        window.mainloop()

def write_to_excel(dt_string,id_sensor,cycle):
    filename = r'C:\Users\amrus\Desktop\project\אקתון\sensorData.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    row = ws.max_row + 1
    ws.cell(column=1, row=row, value=dt_string)
    ws.cell(column=2, row=row, value=id_sensor)
    ws.cell(column=3, row=row, value=cycle)
    wb.save(filename)
def get_from_sentence_parmeters(text):
    text = str(text)
    pos = text.rfind(":")
    num_sensor = text[pos-1]
    if text[pos+2] == '0':
        cycle = text[pos+2:pos+6]
    else:
        cycle = None
    return num_sensor,cycle
def arduino_data():
    readDataPrevies = "123456789"
    ArduinoSerial = serial.Serial('com5', 9600)  # Create Serial port object called arduinoSerialData
    time.sleep(2)  # wait for 2 seconds for the communication to get established

    #get data
    readData = ArduinoSerial.readline()  # read the serial data and print it as line
    num_sensor_prv, cycle_per_sec_prv = get_from_sentence_parmeters(readDataPrevies)
    data_alert =[]
    while(True):
        readDataPrevies = readData
        readData = ArduinoSerial.readline()  # read the serial data and print it as line
        print(readData)
        num_sensor, cycle_per_sec = get_from_sentence_parmeters(readData)
        num_sensor_prv, cycle_per_sec_prv = get_from_sentence_parmeters(readDataPrevies)

        if (cycle_per_sec != None) and ((cycle_per_sec_prv != cycle_per_sec) or (num_sensor != num_sensor_prv)):
            # get time of the invert
            now = datetime.now()
            dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
            if float(cycle_per_sec) < 0.2:
                data_alert.append([dt_string,num_sensor,cycle_per_sec])
                print(data_alert)
                write_to_excel(dt_string, num_sensor, cycle_per_sec)


if __name__ == "__main__":
    global data_alert
    data_alert = []
    x = threading.Thread(target=ux, args=(1,))
    x.start()
    arduino_data()


